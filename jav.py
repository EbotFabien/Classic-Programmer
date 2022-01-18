import openpyxl
import xlrd
loc = ("C:/Users/user/Downloads/Base histo missions EDL 2020-08.xlsx")
#wb = xlrd.open_workbook(loc)
#sheet = wb.sheet_by_index(0)

#sheet.cell_value(0, 0)
wb_obj = openpyxl.load_workbook(loc)
print('ok')
sheet=wb_obj.active
#data ={'M','AP','AS'}
for i in range(1,5):
    print(sheet["AS"][i].value)

#for i, row in enumerate(sheet["M"].iter_rows(values_only=True)):
  #  print(i)
#for column in sheet.iter_cols(1, sheet.max_column):
      #  col_names.append(column[0].value)

#print(col_names)        
#wb = xlrd.open_workbook(loc)
#sheet = wb.sheet_by_index(0)

#sheet.cell_value(0, 0)


#for i in range(0,5):
       # name=sheet.row_values(i+1)
        #print(name[12])
